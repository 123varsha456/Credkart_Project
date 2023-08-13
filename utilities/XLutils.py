import openpyxl


def ReadData(file,sheetname,rownum,columnum):
    workbook=openpyxl.load_workbook(file)  #file open krte
    sheet=workbook[sheetname]          #sheet
    return sheet.cell(row=rownum,column=columnum).value


def WriteData(file , sheetname , rownum , columnum,data):
    workbook=openpyxl.load_workbook(file)   ##file
    sheet=workbook[sheetname]         #sheet
    sheet.cell(row=rownum,column=columnum).value = data   ##data add krto
    workbook.save(file)              #save file


def RowCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)  ##file
    sheet=workbook[sheetname]            #sheet
    return sheet.max_row        #row kiti ahet te return krto

